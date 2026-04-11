"""REIT / ETF のフィルタリング

JPX が公開する上場銘柄一覧 Excel を取得し、
「市場・商品区分」列から ETF/ETN/REIT/インフラファンド等を正確に判定する。
証券コード範囲による近似判定は使用しない。
"""

import os
import requests
import openpyxl
from io import BytesIO

# JPX 上場銘柄一覧 Excel のURL
# ※ JPX は定期的にURLを更新するため、取得失敗時はフォールバック
JPX_LIST_URL = "https://www.jpx.co.jp/markets/statistics-equities/misc/tvdivq0000001vg2-att/data_j.xls"
JPX_LIST_URL_XLSX = "https://www.jpx.co.jp/markets/statistics-equities/misc/tvdivq0000001vg2-att/data_j.xlsx"

# キャッシュファイルパス
CACHE_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "jpx_listed.xlsx")


def _download_jpx_list() -> bytes:
    """JPX 上場銘柄一覧をダウンロードする"""
    for url in [JPX_LIST_URL_XLSX, JPX_LIST_URL]:
        try:
            print(f"  Downloading JPX list from: {url}")
            resp = requests.get(url, timeout=60)
            if resp.status_code == 200 and len(resp.content) > 1000:
                return resp.content
        except requests.RequestException:
            continue
    raise RuntimeError("JPX 上場銘柄一覧のダウンロードに失敗しました")


def get_excluded_codes() -> set[str]:
    """
    REIT / ETF / ETN / インフラファンド 等の証券コードセットを返す。

    JPX Excel の「市場・商品区分」列に以下が含まれるものを除外対象とする:
      - ETF/ETN
      - REIT（不動産投資信託）
      - インフラファンド
      - 出資証券 等

    返り値: 4桁証券コードの set
    """
    content = _download_jpx_list()
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    # ヘッダー行を探す
    header_row = None
    code_col = None
    segment_col = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=False), start=1):
        for cell in row:
            val = str(cell.value or "").strip()
            if "コード" in val and code_col is None:
                code_col = cell.column - 1
                header_row = row_idx
            if "市場・商品区分" in val or "市場商品区分" in val:
                segment_col = cell.column - 1
                header_row = row_idx

    if code_col is None or segment_col is None:
        # フォールバック: 一般的な列位置 (A=コード, C=市場・商品区分)
        print("  Warning: Header detection failed, using default columns")
        code_col = 0
        segment_col = 2
        header_row = 1

    # 除外キーワード
    exclude_keywords = [
        "ETF", "ETN",
        "REIT", "不動産投資信託",
        "インフラファンド", "インフラ投資法人",
        "出資証券",
        "ベンチャーファンド",
    ]

    excluded: set[str] = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[code_col] is None:
            continue
        code = str(row[code_col]).strip()
        # 4桁に正規化
        code = code[:4] if len(code) >= 4 else code

        segment = str(row[segment_col] or "").strip()
        if any(kw in segment for kw in exclude_keywords):
            excluded.add(code)

    wb.close()
    print(f"  Excluded codes (REIT/ETF/etc.): {len(excluded)} companies")
    return excluded


def filter_disclosures(disclosures: list, excluded_codes: set[str]) -> list:
    """REIT/ETF を除外した開示リストを返す"""
    before = len(disclosures)
    filtered = [d for d in disclosures if d.code not in excluded_codes]
    after = len(filtered)
    print(f"  Filtered: {before} -> {after} (removed {before - after} REIT/ETF disclosures)")
    return filtered
